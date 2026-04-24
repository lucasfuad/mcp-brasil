"""DuckDB-based dataset loader and query executor (ADR-004).

Contract:
    - load_into_duckdb() downloads the source CSV and materializes it as a
      table inside a persistent .duckdb file.
    - ensure_loaded() is idempotent: uses the cache if fresh, refreshes if
      expired/missing based on DATASET_REFRESH_MODE.
    - executar_query() opens the DuckDB file in read-only mode and runs a
      parameterized SELECT, returning rows as list[dict].
"""

from __future__ import annotations

import asyncio
import contextlib
import logging
import time
from collections.abc import Iterable, Iterator
from pathlib import Path
from typing import Any

import duckdb

from mcp_brasil import settings

from .cache import (
    Manifest,
    db_path,
    load_manifest,
    save_manifest,
)
from .dataset import DatasetSpec

logger = logging.getLogger(__name__)

# Per-dataset locks prevent concurrent loads of the same dataset
_LOAD_LOCKS: dict[str, asyncio.Lock] = {}


def _lock_for(dataset_id: str) -> asyncio.Lock:
    lock = _LOAD_LOCKS.get(dataset_id)
    if lock is None:
        lock = _LOAD_LOCKS[dataset_id] = asyncio.Lock()
    return lock


def _is_fresh(manifest: Manifest | None, spec: DatasetSpec) -> bool:
    """Return True if the cached data is within TTL."""
    if manifest is None or manifest.fetched_at <= 0:
        return False
    if settings.DATASET_REFRESH_MODE == "force":
        return False
    if settings.DATASET_REFRESH_MODE == "never":
        return True  # trust cache even if stale
    age_days = (time.time() - manifest.fetched_at) / 86400.0
    return age_days < spec.ttl_days


def _render_csv_options(opts: dict[str, Any]) -> str:
    """Serialize duckdb.read_csv_auto options into SQL-safe kwargs."""
    parts: list[str] = []
    for key, val in opts.items():
        if isinstance(val, bool):
            parts.append(f"{key}={str(val).lower()}")
        elif isinstance(val, int | float):
            parts.append(f"{key}={val}")
        elif isinstance(val, str):
            escaped = val.replace("'", "''")
            parts.append(f"{key}='{escaped}'")
        elif isinstance(val, list):
            inner = ",".join(f"'{str(x).replace(chr(39), chr(39) * 2)}'" for x in val)
            parts.append(f"{key}=[{inner}]")
        elif isinstance(val, dict):
            # dtypes / column_types / types — STRUCT-literal form
            struct_parts = []
            for col_name, col_type in val.items():
                col_escaped = str(col_name).replace("'", "''")
                type_escaped = str(col_type).replace("'", "''")
                struct_parts.append(f"'{col_escaped}': '{type_escaped}'")
            parts.append(f"{key}={{{', '.join(struct_parts)}}}")
        else:
            raise TypeError(f"Unsupported CSV option type for {key}: {type(val)}")
    return ", ".join(parts)


_DUCKDB_ENCODINGS = {"utf-8", "utf8", "latin-1", "latin1", "utf-16", "utf16"}


def _transcode_stream(
    read_bytes: Iterable[bytes],
    dest: Path,
    source_encoding: str,
) -> int:
    """Write ``read_bytes`` iterable to ``dest``, transcoding if needed."""
    import codecs

    normalized = source_encoding.lower().replace("_", "-")
    needs_transcode = normalized not in _DUCKDB_ENCODINGS

    total = 0
    with dest.open("wb") as f:
        if not needs_transcode:
            for chunk in read_bytes:
                f.write(chunk)
                total += len(chunk)
        else:
            decoder = codecs.getincrementaldecoder(source_encoding)(errors="replace")
            for chunk in read_bytes:
                text = decoder.decode(chunk)
                if text:
                    encoded = text.encode("utf-8")
                    f.write(encoded)
                    total += len(encoded)
            tail = decoder.decode(b"", final=True)
            if tail:
                encoded = tail.encode("utf-8")
                f.write(encoded)
                total += len(encoded)
    return total


def _download_to_file(
    url: str,
    dest: Path,
    timeout: float,
    source_encoding: str = "utf-8",
) -> int:
    """Stream a remote file to disk via httpx (follows redirects).

    If the declared ``source_encoding`` is not one that DuckDB supports,
    the stream is transcoded to UTF-8 on the fly (common for Windows-1252
    Brazilian gov files with smart quotes/en-dashes).

    Returns size in bytes written to disk.
    """
    import httpx

    with (
        httpx.Client(follow_redirects=True, timeout=timeout) as client,
        client.stream("GET", url) as resp,
    ):
        resp.raise_for_status()
        return _transcode_stream(
            resp.iter_bytes(chunk_size=1_048_576),
            dest,
            source_encoding,
        )


def _download_and_extract_zip(
    url: str,
    dest: Path,
    *,
    zip_member: str,
    timeout: float,
    source_encoding: str = "utf-8",
) -> int:
    """Download a ZIP from ``url``, extract ``zip_member`` to ``dest``.

    The ZIP is first streamed to a temp file (zip central directory is at the
    end, so streaming extraction of a specific member is impractical). Then
    the chosen member is transcoded to UTF-8 if needed.
    """
    import zipfile

    import httpx

    zip_tmp = dest.with_suffix(".zip.tmp")
    logger.info("Downloading ZIP %s to %s", url, zip_tmp)
    with (
        httpx.Client(follow_redirects=True, timeout=timeout) as client,
        client.stream("GET", url) as resp,
        zip_tmp.open("wb") as zf,
    ):
        resp.raise_for_status()
        for chunk in resp.iter_bytes(chunk_size=4_194_304):
            zf.write(chunk)

    logger.info("Extracting %s from ZIP", zip_member)
    with zipfile.ZipFile(zip_tmp) as z:
        names = z.namelist()
        if zip_member not in names:
            # Support simple glob match: exact member preferred, otherwise
            # the first member whose name contains the requested string.
            match = next((n for n in names if zip_member in n), None)
            if match is None:
                raise RuntimeError(
                    f"ZIP at {url!r} has no member matching {zip_member!r}. "
                    f"Available: {names[:5]}..."
                )
            zip_member = match

        with z.open(zip_member) as src:

            def _chunks() -> Iterator[bytes]:
                while True:
                    chunk = src.read(1_048_576)
                    if not chunk:
                        return
                    yield chunk

            total = _transcode_stream(_chunks(), dest, source_encoding)

    with contextlib.suppress(FileNotFoundError):
        zip_tmp.unlink()
    return total


def _download_binary(url: str, dest: Path, timeout: float) -> int:
    """Stream a binary file (Parquet/XLSX) to disk — no transcoding."""
    import httpx

    total = 0
    with (
        httpx.Client(follow_redirects=True, timeout=timeout) as client,
        client.stream("GET", url) as resp,
        dest.open("wb") as f,
    ):
        resp.raise_for_status()
        for chunk in resp.iter_bytes(chunk_size=1_048_576):
            f.write(chunk)
            total += len(chunk)
    return total


def _xlsx_to_csv(xlsx_path: Path, csv_path: Path, sheet: str | int = 0) -> int:
    """Convert one sheet of an XLSX file to a UTF-8 CSV at ``csv_path``.

    Uses openpyxl in read-only mode (streaming) so big files don't blow
    memory. Returns bytes written.
    """
    import csv as _csv

    from openpyxl import load_workbook  # type: ignore[import-untyped]

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if isinstance(sheet, int):
            names = wb.sheetnames
            if sheet < 0 or sheet >= len(names):
                raise RuntimeError(
                    f"XLSX sheet index {sheet} out of range (have {len(names)} sheets)"
                )
            ws = wb[names[sheet]]
        else:
            if sheet not in wb.sheetnames:
                raise RuntimeError(
                    f"XLSX has no sheet named {sheet!r}. Available: {wb.sheetnames}"
                )
            ws = wb[sheet]

        total = 0
        with csv_path.open("w", encoding="utf-8", newline="") as f:
            writer = _csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if v is None else v for v in row])
            total = f.tell()
        return total
    finally:
        wb.close()


def _stage_source(
    spec: DatasetSpec,
    url: str,
    zip_member: str | None,
    dest: Path,
    timeout: float,
) -> int:
    """Download (+ extract/convert) a single source to ``dest``.

    CSV → UTF-8 (transcoded if needed).
    XLSX → downloaded then converted to UTF-8 CSV in-place.
    Parquet → downloaded as-is (binary, no transcoding).
    """
    if spec.source_format == "parquet":
        return _download_binary(url, dest, timeout)
    if spec.source_format == "xlsx":
        # openpyxl validates the extension — keep .xlsx literally.
        xlsx_tmp = dest.parent / f"{dest.stem}.xlsx"
        _download_binary(url, xlsx_tmp, timeout)
        try:
            return _xlsx_to_csv(xlsx_tmp, dest, sheet=spec.xlsx_sheet)
        finally:
            with contextlib.suppress(FileNotFoundError):
                xlsx_tmp.unlink()
    if zip_member:
        return _download_and_extract_zip(
            url,
            dest,
            zip_member=zip_member,
            timeout=timeout,
            source_encoding=spec.source_encoding,
        )
    return _download_to_file(url, dest, timeout=timeout, source_encoding=spec.source_encoding)


def _render_read_call(staged_path: Path, spec: DatasetSpec) -> str:
    """Render a DuckDB read call string for the staged source file.

    Dispatches by ``spec.source_format``:
        - parquet → ``read_parquet('<path>')`` (csv_options ignored)
        - csv/xlsx → ``read_csv_auto('<path>', ...)`` with csv_options
    """
    path_escaped = str(staged_path).replace("'", "''")
    if spec.source_format == "parquet":
        return f"read_parquet('{path_escaped}')"
    kwargs = _render_csv_options({**spec.csv_options, "encoding": "utf-8"})
    kwargs_part = f", {kwargs}" if kwargs else ""
    return f"read_csv_auto('{path_escaped}'{kwargs_part})"


def _load_into_duckdb(spec: DatasetSpec) -> Manifest:
    """Download the source(s) and materialize them in a persistent .duckdb file.

    Strategy:
        1. Stream each CSV to a local temp file via httpx (reliable with
           redirects/cookies/slow servers; handles ZIP archives via
           ``zip_member``; transcodes cp1252/other encodings to UTF-8).
        2. Load into DuckDB in an **ephemeral** tempdir (DuckDB uses mmap
           which does not work on SMB-backed volumes like Azure Files).
        3. After ingestion, atomically move the finished .duckdb file to
           its final location in the cache dir (which may be SMB-mounted).

    For single-source specs, the table is created directly. For multi-source
    specs, one table is created per source (named ``{table}_{suffix}``) plus
    a ``{table}`` VIEW doing ``UNION ALL BY NAME`` across them.

    Runs synchronously — caller is responsible for off-loading to a thread
    via ``asyncio.to_thread`` when called from async context.
    """
    import shutil
    import tempfile

    from mcp_brasil import settings as _settings

    path = db_path(spec.id)
    # Write DuckDB file to ephemeral tempdir; mmap/fcntl on SMB shares
    # (Azure Files) does not support all the operations DuckDB needs.
    ephemeral_dir = Path(tempfile.mkdtemp(prefix=f"mcpb-{spec.id}-"))
    tmp_path = ephemeral_dir / f"{spec.id}.duckdb"
    # Clean any leftover .part in the cache (from a previous failed run on
    # an older code path); the new file flow doesn't use it, but residue
    # wastes space.
    old_part = path.with_suffix(".duckdb.part")
    with contextlib.suppress(FileNotFoundError):
        old_part.unlink()

    # Normalize the source list — single-source specs are promoted to a
    # one-entry list so the downstream loop handles both uniformly.
    sources: list[tuple[str, str | None, str]] = (
        list(spec.sources) if spec.sources else [(spec.url, spec.zip_member, "")]
    )

    logger.info(
        "Loading dataset %s with %d source(s) into %s (final: %s)",
        spec.id,
        len(sources),
        tmp_path,
        path,
    )

    def _open_con() -> duckdb.DuckDBPyConnection:
        """Open DuckDB with tight memory caps.

        Each iteration uses a fresh connection so Python GC reclaims
        DuckDB's buffer pool between sources — PRAGMA memory_limit alone
        doesn't release pages held by the arena allocator.
        """
        c = duckdb.connect(str(tmp_path), read_only=False)
        with contextlib.suppress(duckdb.Error):
            c.execute("PRAGMA memory_limit='2GB'")
            c.execute(f"PRAGMA temp_directory='{ephemeral_dir}/duckdb-spill'")
            c.execute("PRAGMA threads=2")
        return c

    total_row_count = 0
    schema_reprs: list[str] = []
    con: duckdb.DuckDBPyConnection | None = None
    try:
        # Extension of the staged file depends on source_format — parquet
        # stays binary; csv/xlsx end up as UTF-8 CSV after _stage_source.
        staged_ext = "parquet" if spec.source_format == "parquet" else "csv"
        for url, zip_member, suffix in sources:
            # Staged temp lives in the ephemeral dir (no SMB churn, fast I/O).
            staged_tmp = ephemeral_dir / f"source-{suffix or 'single'}.{staged_ext}"
            if staged_tmp.exists():
                staged_tmp.unlink()
            staged = _stage_source(
                spec, url, zip_member, staged_tmp, _settings.DATASET_DOWNLOAD_TIMEOUT
            )
            logger.info(
                "Staged source %s (%d bytes); ingesting into DuckDB",
                suffix or spec.table,
                staged,
            )
            table_name = f"{spec.table}_{suffix}" if suffix else spec.table
            read_call = _render_read_call(staged_tmp, spec)

            con = _open_con()
            try:
                con.execute(f'CREATE OR REPLACE TABLE "{table_name}" AS SELECT * FROM {read_call}')
                row = con.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()
                total_row_count += int(row[0]) if row else 0
                schema_rows = con.execute(f'DESCRIBE "{table_name}"').fetchall()
                schema_reprs.append(
                    f"{table_name}:" + "|".join(f"{r[0]}:{r[1]}" for r in schema_rows)
                )
                with contextlib.suppress(duckdb.Error):
                    con.execute("CHECKPOINT")
            finally:
                con.close()
            with contextlib.suppress(FileNotFoundError):
                staged_tmp.unlink()
            logger.info("Ingested %s; connection closed", table_name)

        # Final pass: open once to build the UNION ALL BY NAME view.
        con = _open_con()

        # Multi-source: create consolidated UNION ALL BY NAME view.
        if spec.sources:
            parts = [f'SELECT * FROM "{spec.table}_{sfx}"' for _, _, sfx in spec.sources]
            union_sql = " UNION ALL BY NAME ".join(parts)
            con.execute(f'CREATE OR REPLACE VIEW "{spec.table}" AS {union_sql}')

        schema_repr = "||".join(schema_reprs)
        row_count = total_row_count
    finally:
        if con is not None:
            con.close()

    # Move the completed DuckDB file from ephemeral to the persistent cache.
    # shutil.move handles cross-filesystem (tmpfs -> SMB) transparently.
    with contextlib.suppress(FileNotFoundError):
        path.unlink()
    shutil.move(str(tmp_path), str(path))
    # Clean up the ephemeral workdir
    with contextlib.suppress(FileNotFoundError, OSError):
        shutil.rmtree(ephemeral_dir, ignore_errors=True)
    size_bytes = path.stat().st_size

    manifest = Manifest(
        id=spec.id,
        url=spec.url,
        table=spec.table,
        fetched_at=time.time(),
        row_count=row_count,
        size_bytes=size_bytes,
        schema_hash=_hash_text(schema_repr),
        source=spec.source,
    )
    save_manifest(manifest)
    logger.info("Dataset %s ready: %d rows, %d bytes", spec.id, row_count, size_bytes)
    return manifest


def _hash_text(text: str) -> str:
    """Short, stable hash for schema drift detection."""
    import hashlib

    return hashlib.sha256(text.encode("utf-8")).hexdigest()[:16]


async def ensure_loaded(spec: DatasetSpec) -> Manifest:
    """Idempotent: load the dataset if missing/stale; return current manifest.

    Respects MCP_BRASIL_DATASET_REFRESH:
        - auto (default): refresh when age > ttl_days
        - never: use cache regardless of age; error if cache missing
        - force: always re-download
    """
    lock = _lock_for(spec.id)
    async with lock:
        manifest = load_manifest(spec.id)
        if _is_fresh(manifest, spec) and db_path(spec.id).exists():
            assert manifest is not None
            return manifest
        if settings.DATASET_REFRESH_MODE == "never":
            if manifest is None or not db_path(spec.id).exists():
                raise RuntimeError(f"Dataset {spec.id!r} has no cache and refresh=never")
            return manifest
        # Download in a worker thread so we don't block the event loop.
        return await asyncio.to_thread(_load_into_duckdb, spec)


async def refresh_dataset(spec: DatasetSpec) -> Manifest:
    """Force a re-download and return the new manifest."""
    lock = _lock_for(spec.id)
    async with lock:
        return await asyncio.to_thread(_load_into_duckdb, spec)


async def executar_query(
    spec: DatasetSpec,
    sql: str,
    params: list[Any] | tuple[Any, ...] = (),
) -> list[dict[str, Any]]:
    """Run a parameterized read-only query against a loaded dataset.

    Ensures the dataset is loaded first (may trigger download). Opens
    DuckDB in read-only mode — DDL/DML raise ``duckdb.Error``.
    """
    await ensure_loaded(spec)
    return await asyncio.to_thread(_execute_sync, spec, sql, list(params))


def _execute_sync(spec: DatasetSpec, sql: str, params: list[Any]) -> list[dict[str, Any]]:
    path = db_path(spec.id)
    con = duckdb.connect(str(path), read_only=True)
    try:
        cursor = con.execute(sql, params) if params else con.execute(sql)
        cols = [c[0] for c in cursor.description] if cursor.description else []
        return [dict(zip(cols, row, strict=False)) for row in cursor.fetchall()]
    finally:
        con.close()


async def get_status(spec: DatasetSpec) -> dict[str, Any]:
    """Introspect current cache state for a dataset (no network)."""
    manifest = load_manifest(spec.id)
    path = db_path(spec.id)
    exists = path.exists()
    age_days: float | None = None
    if manifest and manifest.fetched_at > 0:
        age_days = (time.time() - manifest.fetched_at) / 86400.0
    return {
        "id": spec.id,
        "cached": exists,
        "row_count": manifest.row_count if manifest else 0,
        "size_bytes": manifest.size_bytes if manifest else 0,
        "fetched_at": manifest.fetched_at if manifest else 0.0,
        "age_days": age_days,
        "fresh": _is_fresh(manifest, spec),
        "ttl_days": spec.ttl_days,
        "url": spec.url,
        "table": spec.table,
        "source": spec.source,
    }
